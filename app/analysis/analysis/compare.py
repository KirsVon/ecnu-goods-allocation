import json
from openpyxl import load_workbook


def spiral_welded_pipe(itemid):
    """计算螺旋焊管的件重

    Args:
        tiemid:规格

    Returns:
        result: 件重

    Raise:

    """
    # 分割规格数据
    temp1 = itemid.split("*")
    # 外径
    outside = float((temp1[0][3:]).lstrip("0"))
    # 壁厚
    wall = float(temp1[1])
    # 长
    length = float(temp1[2]) / 1000
    # (外径-(壁厚-0.9))*(壁厚-0.9)*0.0246615=米重(公斤)*12米
    result = ((outside - (wall - 0.9)) * (wall - 0.9) * 0.0246615) * length
    return result


def rectangular_pipe(itemid):
    """计算方矩管的件重

    Args:
       itemid: 规格

    Returns:
        result: 件重

    Raise:

    """
    # 分割规格数据
    temp1 = itemid.split("*")
    # 宽
    width = float((temp1[0][3:]).lstrip("0"))
    # 高
    height = float(temp1[1])
    # 壁厚
    wall = float(temp1[2])
    # 长
    length = float(temp1[3]) / 1000
    # 周长
    c = (width + height) * 2
    # ((周长-3)/3.14159-壁厚)*壁厚*0.02466=米重*6米
    result = (((c - 3) / 3.14159 - wall) * wall * 0.02466) * length
    return result


def welded_pipe(itemid):
    """计算焊管的件重

    Args:
        itemid: 规格

    Returns:
        result: 件重

    Raise:

    """
    # 分割规格数据
    temp1 = itemid.split("*")
    # 外径
    outside = float((temp1[0][3:]).lstrip("0"))
    # 壁厚
    wall = float(temp1[1])
    # 长
    length = float(temp1[2]) / 1000
    # (外径-壁厚)*壁厚*0.02466=米重*6米
    result = ((outside - wall) * wall * 0.02466) * length
    return result


def write_xlsx(data):
    """写入xlsx

    Args:
        data: 要写入的数据
    Returns:

    Raise:

    """
    # 加载xlsx表
    wb = load_workbook("test1.xlsx")
    # 激活sheet
    ws = wb.active
    # 获取最大的行数
    row = ws.max_row
    # 得到下一行的行号
    x = row + 1
    # 为下一行添加数据
    ws.cell(x, 1, data[0])
    ws.cell(x, 2, data[1])
    ws.cell(x, 3, data[2])
    ws.cell(x, 4, data[3])
    ws.cell(x, 5, data[4])
    ws.cell(x, 6, data[5])
    # 保存
    wb.save("test1.xlsx")
    # 关闭
    wb.close()


# 打开json文件 读取数据
with open("json.txt", "rt") as f:
    temp = json.loads(f.read())
# 关闭文件
f.close()
# 遍历库存数据
for i in temp["data"]["list"]:
    # 品名
    cname = i["cname"]
    # 规格
    itemid = i["itemid"]
    # 重量
    endqty = float(i["endqty"])
    # 总根数
    zg00 = float(i["zg00"])
    # 仓库
    whsDesc = i["whsDesc"]
    # 库位
    locid = i["locid"]
    # 根据仓库数据计算件重
    if zg00 != 0:
        div_result = endqty / zg00
    else:
        div_result = "总根数为0"
    # 分割规格数据
    list1 = itemid.split("*")
    # 根据不同的品名，通过公式计算出件重result，写入xlsx
    if cname == "焊管":
        result = welded_pipe(itemid)
        write_xlsx([cname, itemid, div_result, result, whsDesc, locid])
    if cname == "螺旋焊管":
        result = spiral_welded_pipe(itemid)
        write_xlsx([cname, itemid, div_result, result, whsDesc, locid])
    if cname == "方矩管" or cname == "热镀方矩管":
        result = rectangular_pipe(itemid)
        write_xlsx([cname, itemid, div_result, result, whsDesc, locid])










